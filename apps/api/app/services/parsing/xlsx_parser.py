"""XLSX parser: extracts cell text using openpyxl. No macros or binary data executed."""
from app.services.parsing.base import MAX_TEXT_BYTES, CURRENT_PARSER_VERSION, ParserResult

PARSER_NAME = "xlsx"


def parse_xlsx(content: bytes) -> ParserResult:
    try:
        import io
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        return ParserResult(
            text="", parser_name=PARSER_NAME, parser_version=CURRENT_PARSER_VERSION,
            error="openpyxl not installed — XLSX parsing is not available",
        )

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows_text = []
        sheet_count = 0
        for sheet in wb.worksheets:
            sheet_count += 1
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    rows_text.append("\t".join(cells))

        text = "\n".join(rows_text)

        if len(text.encode("utf-8")) > MAX_TEXT_BYTES:
            text = text.encode("utf-8")[:MAX_TEXT_BYTES].decode("utf-8", errors="ignore")

        wb.close()
        return ParserResult(
            text=text,
            parser_name=PARSER_NAME,
            parser_version=CURRENT_PARSER_VERSION,
            metadata_json={"sheet_count": sheet_count},
        )
    except Exception as exc:
        return ParserResult(
            text="", parser_name=PARSER_NAME, parser_version=CURRENT_PARSER_VERSION,
            error=f"XLSX parse error: {str(exc)[:500]}",
        )
